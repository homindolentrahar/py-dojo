from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

data = {
    "20.62.0101": {
        "name": "John Doe",
        "semesters": {
            "1": 3.8,
            "2": 3.4,
            "3": 3.2,
            "4": 4.0,
        }
    },
    "20.62.0102": {
        "name": "Jane Roe",
        "semesters": {
            "1": 4.0,
            "2": 4.0,
            "3": 3.9,
            "4": 4.0,
        }
    },
    "20.62.0103": {
        "name": "Brian Olwen",
        "semesters": {
            "1": 3.5,
            "2": 4.0,
            "3": 3.8,
            "4": 4.0,
        }
    },
    "20.62.0104": {
        "name": "Josephine Law",
        "semesters": {
            "1": 3.6,
            "2": 3.3,
            "3": 4.0,
            "4": 4.0,
        }
    },
    "20.62.0105": {
        "name": "Jacquelin Elvis",
        "semesters": {
            "1": 3.4,
            "2": 3.4,
            "3": 2.2,
            "4": 3.0,
        }
    },
}

work_book = Workbook()
second_year_gp = work_book.active
second_year_gp.title = "Second Year GP"

# Title Cell
second_year_gp.merge_cells("A1:F1")
title_cell = second_year_gp.cell(row=1, column=1)
title_cell.value = "Second Year GP"
title_cell.alignment = Alignment(horizontal='center')
title_cell.font = Font(size=12, bold=True)

second_year_gp.insert_rows(2)
# Data

# Header
headers = ["ID", "Name"] + list(data['20.62.0101']['semesters'].keys())
for col in range(1, len(headers) + 1):
    col_letter = get_column_letter(col)
    second_year_gp[col_letter + '3'] = headers[col - 1]

for key in data:
    name = data[key]['name']
    gp_s = data[key]['semesters'].values()
    row = [key, name] + list(gp_s)
    second_year_gp.append(row)

gp_fields = ['C', 'D', 'E', 'F']

second_year_gp.merge_cells(f"A{len(data) + 4}:B{len(data) + 4}")
average_cell = second_year_gp.cell(row=len(data) + 4, column=1)
average_cell.value = "Average"
average_cell.alignment = Alignment(horizontal='center')
average_cell.font = Font(size=12, bold=True)

for field in gp_fields:
    second_year_gp[f"{field}{len(data) + 4}"] = f"=AVERAGE({field}4:{field}{len(data) + 3})"

second_year_gp.merge_cells(f"A{len(data) + 5}:B{len(data) + 5}")
max_cell = second_year_gp.cell(row=len(data) + 5, column=1)
max_cell.value = "MAX"
max_cell.alignment = Alignment(horizontal='center')
max_cell.font = Font(size=12, bold=True)

for field in gp_fields:
    second_year_gp[f"{field}{len(data) + 5}"] = f"=MAX({field}5:{field}{len(data) + 4})"

second_year_gp.merge_cells(f"A{len(data) + 6}:B{len(data) + 6}")
min_cell = second_year_gp.cell(row=len(data) + 6, column=1)
min_cell.value = "MIN"
min_cell.alignment = Alignment(horizontal='center')
min_cell.font = Font(size=12, bold=True)

for field in gp_fields:
    second_year_gp[f"{field}{len(data) + 6}"] = f"=MIN({field}6:{field}{len(data) + 5})"

work_book.save('SecondYearGP.xlsx')
